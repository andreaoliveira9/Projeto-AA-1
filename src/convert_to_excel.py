import os
import json
from openpyxl import Workbook


def sanitize_sheet_name(name):
    """Sanitiza nomes de folhas para evitar erros no Excel."""
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]
    for char in invalid_chars:
        name = name.replace(char, "")
    return name[:31]  # Limita a 31 caracteres, máximo permitido no Excel


def convert_normal_to_excel(input_dir, files, output_file):
    # Cria o workbook Excel
    wb = Workbook()

    # Processa cada ficheiro JSON
    for file_index, file in enumerate(files):
        print(f"Processing {file}...")

        # Lê o conteúdo do ficheiro JSON
        with open(os.path.join(input_dir, file), "r") as f:
            results = json.load(f)

        # Define o nome da folha, sanitizando se necessário
        sheet_name = sanitize_sheet_name(
            file.replace(".json", "").replace("_clique_search", "")
        )
        ws = wb.create_sheet(title=sheet_name) if file_index > 0 else wb.active
        ws.title = sheet_name

        # Cabeçalhos
        headers = [
            "k",
            "Edge Density",
            "Size",
            "Result",
            "Operations Count",
            "Time (s)",
            "Solution Tested",
            "Timed Out",
        ]
        ws.append(headers)

        # Processa os resultados e escreve na planilha
        for k, densities in results.items():
            for edge_density, sizes in densities.items():
                for size, data in sizes.items():
                    row = [k, edge_density, size]
                    if "timed_out" in data:
                        row += [None, None, None, None, "Sim"]
                    else:
                        row += [
                            (
                                ", ".join(map(str, data.get("result", [])))
                                if data.get("result")
                                else None
                            ),
                            data.get("operations_count", None),
                            data.get("time", None),
                            data.get("solution_tested", None),
                            "Não",
                        ]
                    ws.append(row)

    print(f"Saving all results to {output_file}...")
    # Salva o workbook
    wb.save(output_file)
    print("Excel file created successfully!")


def convert_big_to_excel(input_dir, files, output_file):
    # Cria um workbook Excel
    wb = Workbook()

    for file_index, file in enumerate(files):
        print(f"Processing {file}...")

        # Lê o conteúdo do ficheiro JSON
        with open(os.path.join(input_dir, file), "r") as f:
            results = json.load(f)

        # Define o nome da folha, sanitizando se necessário
        sheet_name = sanitize_sheet_name(
            file.replace(".json", "").replace("_clique_search", "")
        )
        ws = wb.create_sheet(title=sheet_name) if file_index > 0 else wb.active
        ws.title = sheet_name

        # Cabeçalhos
        headers = [
            "k",
            "Result",
            "Operations Count",
            "Time (s)",
            "Solution Tested",
        ]
        ws.append(headers)

        # Processa os dados e escreve no Excel
        for k, data in results.items():
            result = (
                ", ".join(map(str, data.get("result", [])))
                if data.get("result")
                else None
            )
            row = [
                k,
                result,
                data.get("operations_count"),
                data.get("time"),
                data.get("solution_tested"),
            ]
            ws.append(row)

    # Salva o workbook
    try:
        wb.save(output_file)
        print(f"Ficheiro Excel criado: {output_file}")
    except Exception as e:
        print(f"Erro ao salvar o ficheiro Excel: {e}")


if __name__ == "__main__":
    # Diretórios de entrada e saída
    input_dir = "../results/json"
    output_file = "../results/results_1_256.xlsx"

    # Lista de ficheiros JSON no diretório de entrada
    files = [
        f
        for f in os.listdir(input_dir)
        if f.endswith(".json")
        and "10000" not in f
        and "20000" not in f
        and "30000" not in f
    ]

    convert_normal_to_excel(input_dir, files, output_file)

    files = [
        f
        for f in os.listdir(input_dir)
        if f.endswith("10000.json")
        or f.endswith("20000.json")
        or f.endswith("30000.json")
    ]

    output_file = "../results/results_10000_20000_30000.xlsx"

    convert_big_to_excel(input_dir, files, output_file)